from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pprint import pprint
import datetime
import os


def process():
    wb = Workbook()
    ws = wb.active

    # * current date to YYYYMMDD
    today = datetime.datetime.now()
    today_str = today.strftime("%Y%m%d")

    # * get current file directory
    current_dir = os.path.dirname(os.path.abspath(__file__))

    dest_filename = f"{current_dir}/data_extraction_{today_str}.xlsx"

    db_client = MongoClient(
        host="127.0.0.1",
        port=27017,
        username="root",
        password="example",
    )

    db = db_client["dev"]
    plants_collection = db["plants"]
    db_data = list(
        plants_collection.find(
            # {"science_name": "Anemone coronaria"},
            {},
            {
                "_id": 0,
                "science_name": 1,
                "heb_name": 1,
                "life_forms": 1,
                "habitats": 1,
                "locations": 1,
                "description": 1,
            },
        ).limit(0)
    )

    life_forms_set = set(y for x in db_data for y in x["life_forms"])
    habitats_set = set(y for x in db_data for y in x["habitats"])
    locations_set = set(y for x in db_data for y in x["locations"])

    # pprint(locations_set)
    # quit()

    excel_data = []
    # * pivot data
    for i in db_data:
        # * pivot data life_forms
        for j in life_forms_set:
            i[j] = "0"
            if j in i["life_forms"]:
                i[j] = "1"
        # * pivot data habitats
        for j in habitats_set:
            i[j] = "0"
            if j in i["habitats"]:
                i[j] = "1"
        # * pivot data locations
        for j in locations_set:
            i[j] = "0"
            if j in i["locations"]:
                i[j] = "1"

        # print(i)
        i.pop("life_forms")
        i.pop("habitats")
        i.pop("locations")
        excel_data.append(i)

    # quit()

    fieldnames = (
        ["science_name", "heb_name", "description"]
        + list(life_forms_set)
        + list(habitats_set)
        + list(locations_set)
    )

    # * set first header
    start_column = 4
    ws.merge_cells(
        start_row=1,
        start_column=start_column,
        end_row=1,
        end_column=4 + len(life_forms_set) - 1,
    )
    ws.merge_cells(
        start_row=1,
        start_column=start_column + len(life_forms_set),
        end_row=1,
        end_column=4 + len(life_forms_set) + len(habitats_set) - 1,
    )
    ws.merge_cells(
        start_row=1,
        start_column=start_column + len(life_forms_set) + len(habitats_set),
        end_row=1,
        end_column=4 + len(life_forms_set) + len(habitats_set) + len(locations_set) - 1,
    )
    ws.cell(row=1, column=4).value = "life_forms"
    ws.cell(row=1, column=4 + len(life_forms_set)).value = "habitats"
    ws.cell(
        row=1, column=4 + len(life_forms_set) + len(habitats_set)
    ).value = "locations"

    # * set second header
    ws.append(fieldnames)

    # * insert data
    for row in excel_data:
        values = (row[k] for k in fieldnames)
        ws.append(values)

    # * set format for description
    for cell in ws["C"]:
        # print(cell.value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # * set general format
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    ws.column_dimensions["C"].width = 100
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["A"].width = 25

    # * save file
    wb.save(filename=dest_filename)


if __name__ == "__main__":
    """base on request https://trello.com/c/l9cYccVp"""
    process()
